import numpy as np
import torch
import argparse
import pandas as pd
import os
from config import Config
from environment import EnergySystemEnv
from agent import CTDESACAgents
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

def round_dataframe_columns(df):
    skip_columns = ['time_step', '累计充放电次数']
    
    for col in df.columns:
        if col not in skip_columns and df[col].dtype in [np.float64, np.float32, float]:
            if col == '储能电池SOC实时':
                df[col] = df[col].round(4)
            elif col == '制氢效率':
                df[col] = df[col].round(2)
            else:
                df[col] = df[col].round(1)
    return df

def format_csv(df):
    order = [
        'time_step',
        '风速ms', '温度℃', '光强Wm²',
        '风机输出有功kW', '光伏输出有功kW', '储能电池充放电kW', '制氢功率kW',
        '本时刻制氢量Nm3', '理论最大制氢量Nm3', '制氢效率',
        '消纳率', '储能电池SOC实时', '累计充放电次数',
        '1号制氢kW', '2号制氢kW', '3号制氢kW', '4号制氢kW'
    ]
    
    df = df[order]
    
    return df

def save_to_excel(df, filename="best.xlsx"):
    writer = pd.ExcelWriter(filename, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='测试结果', index=False)
    
    workbook = writer.book
    worksheet = writer.sheets['测试结果']
    
    header_format = workbook.add_format({
        'bold': True,
        'font_name': 'Times New Roman',
        'font_size': 12,
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
        'text_wrap': True
    })
    
    cell_format = workbook.add_format({
        'font_name': 'Times New Roman',
        'font_size': 12,
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })
    
    column_widths = {
        'time_step': 10,
        '风速ms': 10,
        '温度℃': 10,
        '光强Wm²': 10,
        '风机输出有功kW': 20,
        '光伏输出有功kW': 20,
        '储能电池充放电kW': 20,
        '制氢功率kW': 12,
        '本时刻制氢量Nm3': 22,
        '理论最大制氢量Nm3': 22,
        '制氢效率': 10,
        '消纳率': 10,
        '储能电池SOC实时': 18,
        '累计充放电次数': 16,
        '1号制氢kW': 12,
        '2号制氢kW': 12,
        '3号制氢kW': 12,
        '4号制氢kW': 12
    }
    
    for i, col in enumerate(df.columns):
        width = column_widths.get(col, 15)
        worksheet.set_column(i, i, width)
    
    worksheet.set_row(0, 30)
    
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)
    
    for row_num in range(len(df)):
        for col_num, value in enumerate(df.iloc[row_num]):
            worksheet.write(row_num + 1, col_num, value, cell_format)
    
    writer.close()
    
    print(f"格式化结果已保存至Excel文件: {filename}")
    
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        times_new_roman = Font(name='Times New Roman', size=12)
        centered = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.font = times_new_roman
                cell.alignment = centered
                cell.border = thin_border
        
        wb.save(filename)
    except Exception as e:
        print(f"Excel后处理出错，但文件已保存: {e}")

def test(model_path=".", model_type="avg_efficiency", save_excel=True):
    config = Config()
    
    print(f"\n正在评估模型: {model_path}")
    print(f"模型类型: {model_type}")
    print(f"测试数据路径: {config.ENV_PARAMS['test_data_path']}")
    print(f"文件是否存在: {os.path.exists(config.ENV_PARAMS['test_data_path'])}")
    
    env = EnergySystemEnv(mode='test')
    
    agents = CTDESACAgents(config)
    
    try:
        agents.load_models(model_path)
        print(f"模型加载成功: {model_path}")
    except Exception as e:
        print(f"模型加载失败: {e}")
        return None, None, None
    
    best_round_hydrogen = 0
    best_round_efficiency = 0
    best_round_data = None
    
    print("\n开始运行100轮测试...")
    
    test_results = {
        'round_idx': [],
        'total_hydrogen': [],
        'overall_efficiency': [],
        'avg_consumption_rate': [],
        'total_switches': []
    }
    
    for round_idx in range(100):
        observations = env.reset(hour_idx=0)
        
        round_data = {
            'time_step': [],
            '风速ms': [],
            '温度℃': [],
            '光强Wm²': [],
            '风机输出有功kW': [],
            '光伏输出有功kW': [],
            '储能电池充放电kW': [],
            '制氢功率kW': [],
            '本时刻制氢量Nm3': [],
            '理论最大制氢量Nm3': [],
            '制氢效率': [],
            '消纳率': [],
            '储能电池SOC实时': [],
            '累计充放电次数': [],
            '1号制氢kW': [],
            '2号制氢kW': [],
            '3号制氢kW': [],
            '4号制氢kW': []
        }
        
        total_hydrogen = 0
        total_theoretical_max = 0
        total_switches = 0
        total_consumption_rate = 0
        
        for step in range(config.ENV_PARAMS['time_steps']):
            current_data = env._get_current_weather_data()
            
            actions = agents.select_actions(observations, evaluate=True)
            next_observations, reward, done, info = env.step(actions)
            
            observations = next_observations
            
            h2_produced = info['hydrogen_produced']
            if isinstance(h2_produced, np.ndarray):
                h2_produced = h2_produced.item() if h2_produced.size == 1 else h2_produced[0]
            total_hydrogen += h2_produced
            
            h2_max = info['theoretical_max_hydrogen']
            if isinstance(h2_max, np.ndarray):
                h2_max = h2_max.item() if h2_max.size == 1 else h2_max[0]
            total_theoretical_max += h2_max
            
            h2_efficiency = h2_produced / h2_max if h2_max > 0 else 0
            
            switch_count = info['switch_count']
            if isinstance(switch_count, np.ndarray):
                switch_count = switch_count.item() if switch_count.size == 1 else switch_count[0]
            total_switches += switch_count
            
            consumption_rate = info['consumption_rate']
            if isinstance(consumption_rate, np.ndarray):
                consumption_rate = consumption_rate.item() if consumption_rate.size == 1 else consumption_rate[0]
            total_consumption_rate += consumption_rate
            
            wind_power = info['wind_power']
            if isinstance(wind_power, np.ndarray):
                wind_power = wind_power.item() if wind_power.size == 1 else wind_power[0]
                
            pv_power = info['pv_power']
            if isinstance(pv_power, np.ndarray):
                pv_power = pv_power.item() if pv_power.size == 1 else pv_power[0]
                
            battery_power = info['battery_power']
            if isinstance(battery_power, np.ndarray):
                battery_power = battery_power.item() if battery_power.size == 1 else battery_power[0]
                
            soc = info['soc']
            if isinstance(soc, np.ndarray):
                soc = soc.item() if soc.size == 1 else soc[0]
            
            electrolyzer_powers = []
            for p in info['electrolyzer_powers']:
                if isinstance(p, np.ndarray):
                    electrolyzer_powers.append(p.item() if p.size == 1 else p[0])
                else:
                    electrolyzer_powers.append(p)
            
            total_elec_power = sum(electrolyzer_powers)
            
            round_data['time_step'].append(step)
            round_data['风速ms'].append(float(current_data['wind_speed']))
            round_data['温度℃'].append(float(current_data['temperature']))
            round_data['光强Wm²'].append(float(current_data['light_intensity']))
            round_data['风机输出有功kW'].append(wind_power)
            round_data['光伏输出有功kW'].append(pv_power)
            round_data['储能电池充放电kW'].append(battery_power)
            round_data['制氢功率kW'].append(total_elec_power)
            round_data['本时刻制氢量Nm3'].append(h2_produced)
            round_data['理论最大制氢量Nm3'].append(h2_max)  
            round_data['制氢效率'].append(h2_efficiency)     
            round_data['消纳率'].append(consumption_rate)
            round_data['储能电池SOC实时'].append(soc)
            round_data['累计充放电次数'].append(total_switches)
            
            for i in range(4):
                round_data[f'{i+1}号制氢kW'].append(electrolyzer_powers[i])
            
            if done:
                break
        
        overall_efficiency = total_hydrogen / total_theoretical_max if total_theoretical_max > 0 else 0
        avg_consumption_rate = total_consumption_rate / config.ENV_PARAMS['time_steps']
        
        total_hydrogen = round(total_hydrogen, 2)
        overall_efficiency = round(overall_efficiency, 4)
        avg_consumption_rate = round(avg_consumption_rate, 4)
        
        test_results['round_idx'].append(round_idx)
        test_results['total_hydrogen'].append(total_hydrogen)
        test_results['overall_efficiency'].append(overall_efficiency)
        test_results['avg_consumption_rate'].append(avg_consumption_rate)
        test_results['total_switches'].append(total_switches)
        
        if (round_idx + 1) % 10 == 0 or round_idx == 0:
            print(f"完成测试轮次 {round_idx+1}/100, "
                  f"当前轮次总制氢量: {total_hydrogen:.2f} Nm³, "
                  f"效率: {overall_efficiency:.4f}")
        
        if total_hydrogen > best_round_hydrogen:
            best_round_hydrogen = total_hydrogen
            best_round_efficiency = overall_efficiency
            best_round_data = round_data.copy()
            print(f"发现新的最佳轮次 (轮次 {round_idx+1}/100): "
                 f"总制氢量 = {best_round_hydrogen:.2f} Nm³, 效率 = {best_round_efficiency:.4f}")
    
    df_results = pd.DataFrame(test_results)
    
    avg_hydrogen = df_results['total_hydrogen'].mean()
    std_hydrogen = df_results['total_hydrogen'].std()
    avg_efficiency = df_results['overall_efficiency'].mean()
    std_efficiency = df_results['overall_efficiency'].std()
    avg_consumption = df_results['avg_consumption_rate'].mean()
    avg_switches = df_results['total_switches'].mean()
    
    print("\n========== 测试统计结果 ==========")
    print(f"模型类型: {model_type}")
    print(f"平均制氢量: {avg_hydrogen:.2f} ± {std_hydrogen:.2f} Nm³")
    print(f"平均制氢效率: {avg_efficiency:.4f} ± {std_efficiency:.4f}")
    print(f"平均消纳率: {avg_consumption:.4f}")
    print(f"平均启停次数: {avg_switches:.2f}")
    print(f"最佳制氢量: {best_round_hydrogen:.2f} Nm³")
    print(f"最佳制氢效率: {best_round_efficiency:.4f}")
    
    stats = {
        'model_type': model_type,
        'avg_hydrogen': avg_hydrogen,
        'std_hydrogen': std_hydrogen,
        'avg_efficiency': avg_efficiency,
        'std_efficiency': std_efficiency,
        'avg_consumption': avg_consumption,
        'avg_switches': avg_switches,
        'best_hydrogen': best_round_hydrogen,
        'best_efficiency': best_round_efficiency
    }
    
    best_df = None
    if best_round_data:
        df_best = pd.DataFrame(best_round_data)
        df_best = round_dataframe_columns(df_best)
        df_best = format_csv(df_best)
        
        # 只在需要时保存Excel
        if save_excel:
            save_to_excel(df_best, "best.xlsx")
            print(f"最佳轮次数据已保存至: best.xlsx")
        
        best_df = df_best
    
    # 返回统计结果、所有测试结果和最佳轮次数据
    return stats, df_results, best_df

def compare_models(model_dir):
    print(f"\n比较同一训练中的不同模型: {model_dir}")
    
    model_types = {
        "avg_efficiency": f"{model_dir}/best_model_avg_efficiency",
        "single_efficiency": f"{model_dir}/best_model_single_efficiency",
        "single_hydrogen": f"{model_dir}/best_model_single_hydrogen",
        "final": f"{model_dir}/final"
    }
    
    global_best_hydrogen = 0
    global_best_model_type = None
    global_best_data = None  # 保存全局最佳数据
    
    for model_type, model_path in model_types.items():
        if not os.path.exists(model_path):
            print(f"警告: 模型路径不存在: {model_path}")
            continue
            
        print(f"\n测试模型类型: {model_type}, 路径: {model_path}")
        
        # 正确解包三个返回值，并禁止中间结果保存Excel
        stats, _, best_data = test(model_path=model_path, model_type=model_type, save_excel=False)
        
        if stats and stats['best_hydrogen'] > global_best_hydrogen:
            global_best_hydrogen = stats['best_hydrogen']
            global_best_model_type = model_type
            global_best_data = best_data  # 保存最佳数据
    
    # 只保存全局最佳模型的数据
    if global_best_data is not None:
        save_to_excel(global_best_data, "best.xlsx")
        print(f"\n全局最佳模型: {global_best_model_type}")
        print(f"最佳制氢量: {global_best_hydrogen:.2f} Nm³")
        print(f"最佳数据已保存至: best.xlsx")
    else:
        print(f"\n未找到有效模型数据")
        
    return global_best_model_type, global_best_hydrogen


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Test RL agents')
    parser.add_argument('--model_path', type=str, default=".", help='Path to saved model directory')
    
    args = parser.parse_args()
    
    compare_models(args.model_path)